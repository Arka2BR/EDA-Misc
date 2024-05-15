#############################################################################################################
#	AUTHOR			: Arkaprabho Adhikary (aadhika1)																	
#	LAST MODIFIED	: 05/29/2023																				
#	DESCRIPTION		: Extracts Z info. for SNPS data.     
#############################################################################################################
import os
from collections import OrderedDict
import pandas as pd
import logging
import json
from glob import glob

import xlsxwriter
import openpyxl
from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.styles import Alignment

#############################################################################################################
#	Main Variables																	        
#############################################################################################################
# The base file should always be a CSV with Refname, Count on the 1st Row
# You can now supply multiple library paths used in the design!
# Example: Refname, Count
#           k0m????? 345
# That should be the format of the input file.
# Output should be .xlsx not .xls as its a redundant and unsupported format.
BASE_FILE	= "/nfs/pdx/stod/stod5017/w.kumardh1.429/experiments/qik.external-ip-neoverse.enyo_core.qik_route_enyo.enyo_core_pdk0.9_I0S_TFM_sweep_run_ww45.6.Expt8_0.55034_IOM_2.42_3.80_1/adept.analysis/analysis.route_opt.design_data.json"
LIBPATH		= ["/nfs/site/disks/ppac_data_zsc9_001/tech_packages/1278/1278.3/PPAC_tech_p1278.3_160h_m14@23ww44.1_pdk783_r0.9_23ww40.5_beta_14ML_i0s_fullset_SUP2FULL/stdcells", "/nfs/site/disks/ppac_data_zsc9_001/tech_packages/1278/1278.3/PPAC_tech_p1278.3_160h_m14@23ww44.1_pdk783_r0.9_23ww40.5_beta_14ML_i0s_fullset_SUP2FULL/stdcells_sup2"]
OUTPUT_FILE	= "/nfs/pdx/stod/stod5017/w.kumardh1.142/arka_script_env/PDK0.9_I0S.xlsx"

#############################################################################################################
#	Tech Node specific Variables																	        
#############################################################################################################

Z_WIDTH_MAP = {'z1': '1', 'z2': '2' , 'z3': '3'}
LIB_NAME    = 'i0s'

#############################################################################################################
#	Parser																	     
#############################################################################################################

class cell_z_parser:

	def __init__(self, output_file, z_width_map, lib_name):
		self.BLACKLIST		= ['tihi', 'tilo'] #Add odd cells here
		self.z_width_dict	= z_width_map
		self.LIB_NAME		= lib_name
		self.vt_z_dict		= {'hvt':{'Z1':0, 'Z2':0, 'Z3':0},
								'svt':{'Z1':0, 'Z2':0, 'Z3':0},
								'lvt':{'Z1':0, 'Z2':0, 'Z3':0},
								'ulvt':{'Z1':0, 'Z2':0, 'Z3':0},
								'Total':{'Z1':0, 'Z2':0, 'Z3':0}
								}
		self.OUTPUT_FILE	= output_file

		self.hvt_data		= list()
		self.svt_data		= list()
		self.lvt_data		= list()
		self.ulvt_data		= list()

	def lib_z_breakdown(self, cell_list):
		
		self.Tot_Z1, self.Tot_Z2, self.Tot_Z3 = 0, 0, 0

		for cell in cell_list:
			#print(cell[3:7],self.BLACKLIST)
			if not cell[3:7] in self.BLACKLIST:
				Z1, Z2, Z3 	= 0, 0, 0			
	
				cell_vt		= cell[-7]
				#print(cell_vt)
				
				if cell_vt == 'a': lib_series = self.ulvt_data
				elif cell_vt == 'b': lib_series = self.lvt_data
				elif cell_vt == 'c': lib_series = self.svt_data
				elif cell_vt == 'd': lib_series = self.hvt_data
				
				START_FLAG 	= ".subckt {}".format(cell)
				END_FLAG  	= ".ends {}".format(cell)
				
				#print(cell, START_FLAG )
				
				for lib_ver in lib_series:
					for line in lib_ver:
						if START_FLAG in line:
							lib_data = lib_ver

				for i, line in enumerate(lib_data):
					#print(line)
					if line.startswith(START_FLAG) :
						START_IDX = i
					if line.startswith(END_FLAG):
						END_IDX = i
				
				
				for line in lib_data[START_IDX+1 : END_IDX]:
					if not line.startswith("*"):
						line_lst 	= line.split()
						print(line_lst)
						if line_lst:			
							mono_Z	 	= line_lst[-4].split("=")[-1]
							mult		= int(line_lst[-2].split("=")[-1])
							finger		= int(line_lst[-1].split("=")[-1])
							if mono_Z == self.z_width_dict['z1']:
								Z1 += mult*finger
							elif mono_Z == self.z_width_dict['z2']:
								Z2 += mult*finger
							elif mono_Z == self.z_width_dict['z3']:
								Z3 += mult*finger
				
				Z1_cnt								= Z1 * self.cell_dict[cell]["Count"]
				Z2_cnt								= Z2 * self.cell_dict[cell]["Count"]
				Z3_cnt								= Z3 * self.cell_dict[cell]["Count"]
				
				self.Tot_Z1							+= Z1_cnt
				self.Tot_Z2							+= Z2_cnt
				self.Tot_Z3							+= Z3_cnt
				
				# Data for in depth cell summary page
				self.cell_dict[cell]["unit-Z"] 		= "{}/{}/{}".format(Z1, Z2, Z3)
				self.cell_dict[cell]["Z1 Total"]	= Z1_cnt
				self.cell_dict[cell]["Z2 Total"]	= Z2_cnt
				self.cell_dict[cell]["Z3 Total"]	= Z3_cnt
				
				# Data for general Summary Page
				if cell_vt == 'a': 
					self.vt_z_dict['ulvt']['Z1']	+= Z1_cnt
					self.vt_z_dict['ulvt']['Z2']	+= Z2_cnt
					self.vt_z_dict['ulvt']['Z3']	+= Z3_cnt
				elif cell_vt == 'b':
					self.vt_z_dict['lvt']['Z1']		+= Z1_cnt
					self.vt_z_dict['lvt']['Z2']		+= Z2_cnt
					self.vt_z_dict['lvt']['Z3']		+= Z3_cnt
				elif cell_vt == 'c': 
					self.vt_z_dict['svt']['Z1']		+= Z1_cnt
					self.vt_z_dict['svt']['Z2']		+= Z2_cnt
					self.vt_z_dict['svt']['Z3']		+= Z3_cnt
				elif cell_vt == 'd':
					self.vt_z_dict['hvt']['Z1']		+= Z1_cnt
					self.vt_z_dict['hvt']['Z2']		+= Z2_cnt
					self.vt_z_dict['hvt']['Z3']		+= Z3_cnt
					
				self.vt_z_dict['Total']["Z1"]		= self.Tot_Z1
				self.vt_z_dict['Total']["Z2"]		= self.Tot_Z2
				self.vt_z_dict['Total']["Z3"]		= self.Tot_Z3

	def lib_data_extractor(self, libpath):
		with open(libpath, 'r') as f:
			return f.readlines()
			
	def workbook_gen(self, filepath, dataframe, sheetname):
		#print(filepath)
		if not os.path.exists(filepath):
			ExcelBook = openpyxl.Workbook()
			ExcelBook.save(filepath)
		else:
			ExcelBook = load_workbook(filepath)
			
		if sheetname in ExcelBook.sheetnames:
			#print(ExcelBook.sheetnames)
			ExcelBook.remove(ExcelBook[sheetname])
		
		with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
			writer.book = ExcelBook
			dataframe.to_excel(writer, sheet_name= sheetname)
		logging.info("\u001b[35;1m{} \u001b[30;1mgenerated in \u001b[35;1m{} successfully. \u001b[0m".format(sheetname, filepath))
	
	#Currently hard-coded SPF Paths
	#This assumes the sp files are name {Lib General Name (i0s, i0m, k0m, k0t...)}_{vt_type}.sp
	def lib_preprocess(self):
		for libpath in self.libset:
			self.hvt_lib 	= glob(os.path.join(libpath, 'spice/*_hvt.sp'))[0]
			self.hvt_data.append(self.lib_data_extractor(self.hvt_lib))
			self.svt_lib 	= glob(os.path.join(libpath, 'spice/*_svt.sp'))[0]
			self.svt_data.append(self.lib_data_extractor(self.svt_lib))
			self.lvt_lib 	= glob(os.path.join(libpath, 'spice/*_lvt.sp'))[0]
			self.lvt_data.append(self.lib_data_extractor(self.lvt_lib))
			self.ulvt_lib 	= glob(os.path.join(libpath, 'spice/*_ulvt.sp'))[0]
			self.ulvt_data.append(self.lib_data_extractor(self.ulvt_lib))
		
	def Z_nexus(self, basefile, libset):
		self.libset 	= libset
		self.cell_dict	= {}
	
		self.lib_preprocess()
        #Mapping Design Data into Dictionary
		with open(basefile) as f:
			data = json.load(f)
		
		cell_nexus 		= data['data']['cells']
		#all_cell_lst	= [x['name'] for x in cell_nexus]
		for cell in cell_nexus:
			name 	= cell['name']
			count	= cell['count']
			self.cell_dict[name] = {"Count":count}
		
		all_cell_lst	= list(self.cell_dict) #Get all the cell names from the csv
		self.lib_z_breakdown(all_cell_lst)
		
		# Arranging the data
		main_df = pd.DataFrame.from_dict(self.cell_dict, orient = 'index')
		main_df.index.name = 'Ref-Name'
		sum_df	= pd.DataFrame.from_dict(self.vt_z_dict, orient = 'index')
		sum_df.index.name = 'Ref'
		
		#Generate File
		self.workbook_gen(self.OUTPUT_FILE, main_df, 'Cell Z Dist.')
		self.workbook_gen(self.OUTPUT_FILE, sum_df, 'Z Summary')
		print("\u001b[32;1mGenerated file : \u001b[35;1m{} \u001b[0m".format(self.OUTPUT_FILE))
		
#############################################################################################################
# Execution																	     
#############################################################################################################	
if __name__ == "__main__":	
	czp 	= cell_z_parser(OUTPUT_FILE, Z_WIDTH_MAP, LIB_NAME)
	czp.Z_nexus(BASE_FILE, LIBPATH)