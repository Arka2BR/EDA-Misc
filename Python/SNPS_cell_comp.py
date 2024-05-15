import os
from os.path import isdir
from os.path import exists as oPe
from datetime import datetime
from math import ceil
import re
import gzip
import pandas as pd
import logging
import yaml
from collections import OrderedDict as OD
import xlsxwriter
import openpyxl
from openpyxl import Workbook  
from openpyxl import load_workbook
from openpyxl.styles import Alignment


CELL_LISTS	= ["/nfs/site/disks/w.aadhika1.894/Manu/cell_list1.csv", "/nfs/site/disks/w.aadhika1.894/Manu/cell_list2.csv"]

HEADERS		= ["m16 stack", "m14 stack"]

OUTPUT_PATH	= "/nfs/site/disks/w.aadhika1.894/Manu/cell_comp.xlsx"

class cellParser():
	def __init__(self, pathList, headers, output):
		self.paths 		= pathList
		self.headers	= headers
		self.output		= output
		
	def cell_init(self):
		# Buffers and Inverters
		self.BUFFER			= 0
		self.INVERTER		= 0
			
		self.BUFF_AREA		= 0
		self.INV_AREA		= 0
			
		# Mappings
		self.vt_map 		= {'d':'hvt', 'c':'svt', 'b':'lvt', 'a':'ulvt'}
		self.phys_lst		= ['zdsd', 'zh50ba']
		self.cell_vt_init 	= {"ulvt":0, "lvt":0, "svt":0, "hvt":0, "buffinv":0, "buffinv_area":0, "physical":0, "physical_area":0, "total":0, "total_area":0}
		self.BUFF_IDS		= ["bfn", "bfm", "bff",  "cbf"]
		self.INV_IDS		= ["inv"]
		self.TRI_LIBS		= ["ibs", "ibm", "i0s", "i0m", "g1m", "k0t", "k0m", "k0s", "k0d"]

	
	def drive_chronicler(self, storage, drive, count, lookup):
	
		# General variables
		key			= drive
		#print(lookup.match(key))
		
		# Data Storage
		#if key not in storage and (lookup.match(key) is not None):
		if key not in storage:
			storage[key] 			= [count]
		elif (lookup.match(key) is not None):
			storage[key][0] 	+= count
			#storage[key][1] 	+= area
			
		return storage
		
	def leg_chronicler(self, storage, drive, count, lookup):
		if lookup.match(drive) is not None :
		
			# General vars
			leg_name	= ceil(float(drive[:2])/3)
			key			= ft.LEG_HOLDER.format(leg_name)
			
			# Leg Dict related
			if key not in storage :
				storage[key]	= count
			else:
				storage[key]	+= count
				
		return storage
	
	def family_chronicler(self, storage, family, count):
		# General vars
		key			= family
		# Storage		
		if key not in storage:
			storage[key] 		= [count]
		else:
			storage[key][0] 	+= count
			#storage[key][1] 	+= area
			
		return storage
	
	def fam_vs_drive_chronicler(self, storage, family, drive, count):
		# General variables
		t0_key	= family
		t1_key	= drive
		# Data Storage
		if t0_key not in storage :
			storage[t0_key]			= {}
			storage[t0_key][t1_key] = [count]
		elif t1_key not in storage[t0_key]:
			storage[t0_key][t1_key] = [count]
		else:
			storage[t0_key][t1_key][0] 	+= count
			#storage[t0_key][t1_key][1] 	+= area
			
		return storage
	
	def cell_parser(self, data, type):
		
		# Initialize function variables
		self.cell_init()
		
		cell_dict			= {}
		self.STD_CELL_COUNT	= 0
		self.PHYS_COUNT		= 0
		
		physCell_ids		= "zdsd"
		#print(physCell_ids)
		
		for line in data[1:]:
			# Filtering out essential data
			lst 		= line.strip().split(",")		
			cfn 		= lst[0] #Cell Full name
			cell_base	= cfn[3: 7]
			#print(cell_base)
			tech		= self.TRI_LIBS
					
			if cfn[:3] in tech:
				FAMILY		= cfn[3:-8]
			else:
				FAMILY		= cfn[:-4]
			
			DRIVE		= cfn[-4:]		
			inst		= int(lst[1])
			form       	= re.compile(r'^\d+x\d+')
			
			# Counting functional & physical Cells
			if cell_base not in physCell_ids:
				self.STD_CELL_COUNT += inst
			else:
				self.PHYS_COUNT += inst
				
			
			# Case-wise functionality
			if type == "DRIVE":
				cell_dict = self.drive_chronicler(cell_dict, DRIVE, inst, form)
			
			elif type == "LEG" :
				cell_dict = self.leg_chronicler(cell_dict, DRIVE, inst, form)
				
			elif type == "FAMILY":
				cell_dict = self.family_chronicler(cell_dict, FAMILY, inst)
		
			elif type == "BUFFINV":
				if cfn[3:6] in self.BUFF_IDS or FAMILY in self.INV_IDS:
					key 			= cfn
					cell_dict[key] 	= [inst]
				
			elif type == "MAIN":
				cell_dict = self.fam_vs_drive_chronicler(cell_dict, FAMILY, DRIVE, inst)
					
			elif type == "VT_DIST":
				self.netlist_chronicler(cfn, inst, area)
				cell_dict = self.cell_vt_init
			else: 
				print(ft.INCORR_CELL_OPT)
		
		#print(cell_dict)
		return cell_dict
	
	def fileParser(self, path):
		with open(path, 'r') as f:
			raw_data = f.readlines()
		
		return raw_data
		
	def dict_warp_cell(self, cell_csv_dict, format=1):
	
		""" Modify the cell dict"""
		outer_keys_cell	= list(cell_csv_dict.keys())
		mod_cell_dict   = {}
		keys_cell 		= []
		for okey in outer_keys_cell:
			keys 		= [k for k in cell_csv_dict[okey].keys() if k not in keys_cell]
			keys_cell   += keys
			
		if format == 1:		
			for key in keys_cell :
				val_list			= []
				for okey in outer_keys_cell:
				
					if key in cell_csv_dict[okey].keys():
						val_list.append(",".join(str(x) for x in cell_csv_dict[okey][key]))
					else:
						val_list.append("0, 0")
					
				mod_cell_dict[key] = val_list
		elif format == 2:
			for key in keys_cell :
	
				for okey in outer_keys_cell:
				
					INSTANCE_IDX	= "Instance {}".format(okey)
					#AREA_IDX		= "Area {}".format(okey)
					inner_keys		= list(cell_csv_dict[okey].keys())
			
					if key in inner_keys:
						if key not in mod_cell_dict:
							mod_cell_dict[key] = {}
							mod_cell_dict[key][INSTANCE_IDX] 	= cell_csv_dict[okey][key][0] 
							#mod_cell_dict[key][AREA_IDX] 		= cell_csv_dict[okey][key][1]
						else:
							mod_cell_dict[key][INSTANCE_IDX] 	= cell_csv_dict[okey][key][0] 
							#mod_cell_dict[key][AREA_IDX] 		= cell_csv_dict[okey][key][1]
					else:
						if key not in mod_cell_dict:
							mod_cell_dict[key] = {}
							mod_cell_dict[key][INSTANCE_IDX] 	= 0
							#mod_cell_dict[key][AREA_IDX] 		= 0
						else:
							mod_cell_dict[key][INSTANCE_IDX] 	= 0
							#mod_cell_dict[key][AREA_IDX] 		= 0
		elif format == 3:
					
			for okey in cell_csv_dict:
				INSTANCE_IDX	= "Instance {}".format(okey)
				#AREA_IDX		= "Area {}".format(okey)
				keys_cell = cell_csv_dict[okey].keys()
				for key in keys_cell:
					print(cell_csv_dict[okey])
					in_keys_cell = 	cell_csv_dict[okey][key]
					if key not in mod_cell_dict:
						mod_cell_dict[key] = {}
					for ik in in_keys_cell:
						print(cell_csv_dict[okey][key])
						if ik not in mod_cell_dict[key].keys():
							mod_cell_dict[key][ik] 					= {INSTANCE_IDX: 0}
							mod_cell_dict[key][ik][INSTANCE_IDX] 	= cell_csv_dict[okey][key]
							#mod_cell_dict[key][ik][AREA_IDX] 		= cell_csv_dict[okey][key][ik][1]
						else:
							mod_cell_dict[key][ik][INSTANCE_IDX] 	= cell_csv_dict[okey][key]
							#mod_cell_dict[key][ik][AREA_IDX] 		= cell_csv_dict[okey][key][ik][1]
							
		#Sorting keys inside the 2nd level	
		for key in mod_cell_dict:
			sub_dict			= mod_cell_dict[key]
			mod_cell_dict[key]	= OD(sorted(sub_dict.items()))	
				
		return 	OD(sorted(mod_cell_dict.items()))
	
	def workbook_gen(self, filepath, dataframe, sheetname):
		#print(filepath)
		if not os.path.exists(filepath):
			ExcelBook = openpyxl.Workbook()
			ExcelBook.save(filepath)
		else:
			ExcelBook = load_workbook(filepath)
			
		if sheetname in ExcelBook.sheetnames:
			#print(ExcelBook.sheetnames)
			ExcelBook.remove(ExcelBook[sheetname])
		
		with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
			writer.book = ExcelBook
			dataframe.to_excel(writer, sheet_name= sheetname)
		logging.info("\u001b[35;1m{} \u001b[30;1mgenerated in \u001b[35;1m{} successfully. \u001b[0m".format(sheetname, filepath))
	
	def cell_csv_gen(self, filepath, data, index_name, sheetname, format = 1):
		if format == 1:
			cell_df = pd.DataFrame.from_dict(data, orient='index')
			#Adding name to index column
			cell_df.index.name = index_name
			# Sorting based on columns
			cell_df.sort_index( axis = 1, ascending = False, inplace=True )
		elif format == 2:
			okeys		= list(data)
			mkeys		= list(data[okeys[0]])
			ikeys		= list(data[okeys[0]][mkeys[0]])
	
			cell_col	= index_name + ikeys
			
			#Multi-index dataframe with tuple of dict keys. Columns made using reset_index
			cell_df =  pd.concat({k: pd.DataFrame(v).T for k, v in data.items()}, axis=0).reset_index() 
			
			#Add index
			cell_df.columns = cell_col
			
			#Rearranging columns
			data_cols	= cell_df.loc[:, cell_df.columns[2:]]
			param_cols	= cell_df.loc[:, cell_df.columns[:2]]
			data_cols.sort_index( axis = 1, ascending = False, inplace=True )
			cell_df		= pd.concat([param_cols, data_cols], axis=1, join = 'inner')
			
			#Removing repeating elements in 1st column
			col_0		= cell_df.columns[0]
			cell_df.loc[cell_df[col_0].duplicated(), col_0] = ''	
			
			#Setting index column
			cell_df		= cell_df.set_index(col_0)
			#print(cell_df)
			
		cell_df 	= cell_df.fillna(0)
		self.workbook_gen(filepath, cell_df, sheetname)		
	
	def cellComp(self):
		comp_drive_dict 	= {}
		comp_fam_dict		= {}
		comp_bufinv_dict	= {}
		comp_main_dict		= {}
		for x, y in zip(self.paths, self.headers):
			data				= self.fileParser(x)
			comp_drive_dict[y] 	= self.cell_parser(data, "DRIVE")
			comp_fam_dict[y]	= self.cell_parser(data, "FAMILY")
			comp_bufinv_dict[y]	= self.cell_parser(data, "BUFFINV")
			comp_main_dict		= self.cell_parser(data, "MAIN")
			
		print(comp_main_dict)
			
		
		comp_drive_dict 	= self.dict_warp_cell(comp_drive_dict, 2)
		comp_fam_dict		= self.dict_warp_cell(comp_fam_dict, 2)
		comp_bufinv_dict	= self.dict_warp_cell(comp_bufinv_dict, 2)
		#comp_main_dict		= self.dict_warp_cell(comp_main_dict, 3)
		
		# Workbook
		workbook 			= self.output
		
		# Cell Tabs
		self.cell_csv_gen(workbook, comp_bufinv_dict, "Buffer-Inverters", "Buffer-Inverters")	
		self.cell_csv_gen(workbook, comp_fam_dict, "Cell Family", "Cell Family")
		self.cell_csv_gen(workbook, comp_drive_dict, "Cell Drive", "Cell Drive")
		#self.cell_csv_gen(workbook, comp_main_dict, ["Family", "Drive"], "Family vs. Drive", 2)
		
		print("\u001b[30;1mWorkbook \u001b[35;1m{} \u001b[30;1mgenerated successfully. \u001b[0m".format(workbook))

if __name__ == "__main__":
	# The 2 ()s is necessary and not a typo!
	cp = cellParser(CELL_LISTS, HEADERS, OUTPUT_PATH)
	cp.cellComp()